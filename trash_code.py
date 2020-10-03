'''
This file is convenient for our development.
when we finish this development, we will delete it,
so,  don't care about it.
'''


import openpyxl
import os
import pydicom

'''

def read_truth_table(case_number):
    file_path = os.path.split(os.path.realpath(__file__))[0] + "/truth_table_path"
    file = open(file_path)
    for line in file:
        wb = openpyxl.load_workbook(line)
        sheet = wb['Sheet1']
        truth_datas = {}
        colName =[]
        for colNum in range(0, sheet.max_column):
            colName.append(sheet[1][colNum].value)
        # print(colName)

        for rowNum in range(2, sheet.max_row):   # skip the first row
            truthcase = {}
            for colNum in range(1, sheet.max_column):
                truthcase[colName[colNum-1]] = str(sheet[rowNum][colNum-1].value)
            truth_datas[rowNum - 1] = truthcase
        return truth_datas[case_number]

truth_datas = read_truth_table(6)
print(truth_datas['gantry'])
'''

#
# for index in truth_datas:
#     print(truth_datas[index])


print("-----------------------------------------------------------")
# print(rows[0])  # 获取第一行
# print(rows[0][0])  # 获取第一行第一列的单元格对象


# file_path = "/Users/yaozhiyuan/myunimelb/semster3/software project/DICOM/LII DICOM samples/YellowLvlIII_7a.dcm"
# res = extract_gantry_angle(file_path)
# print(res)


# Beam Dose Point SSD--    Referenced Dose Reference Sequence--
# Control Point Sequence --   Beam Sequence


def extract_gantry_angle(file_path):
    # step1: read one dicom file from given path and assigns value to the variable ds
    try:
        ds = pydicom.dcmread(file_path, force=True)
    except IOError:
        print("Error: The file was not found or failed to read")
    res = []
    # step2: find all gantry angles from ds
    # gantry angle is located in dicomData.BeamSequence.Item_1.ControlPointSequence.Item_1.GantryAngle
    try:
        for bs in ds.BeamSequence:
            for cp in bs.ControlPointSequence:
                if hasattr(cp, 'GantryAngle'):
                    res.append(cp.GantryAngle)
    except AttributeError as err:
        print("OS error: {0}".format(err))

    # step3: Use 'set' to remove the same value
    res = list(set(res))
    return res

#  2.SSD? why this value is not same format with the truth value
def extract_SSD(file_path):
    # step1: read one dicom file from given path and assigns value to the variable ds
    try:
        ds = pydicom.dcmread(file_path, force=True)
    except IOError:
        print("Error: The file was not found or failed to read")
    res = []
    print(ds)
    # step2. find all gantry angles from ds
    # SSD:
    # Source to Surface Distance
    # Control Point Sequence
    try:
        for bs in ds.BeamSequence:
            for cp in bs.ControlPointSequence:
                if hasattr(cp, 'SourceToSurfaceDistance'):
                            res.append(cp.SourceToSurfaceDistance)
    except AttributeError as err:
        print("OS error: {0}".format(err))

    # step3: Use 'set' to remove the same value
    res = list(set(res))
    return res


# file_path = "/Users/yaozhiyuan/myunimelb/semster3/software project/DICOM/LII DICOM samples/ACDSLevelII3_2a.dcm"

file_path = "/Users/yaozhiyuan/myunimelb/semster3/software project/DICOM/LII DICOM samples/YellowLvlIII_7a.dcm"

res = extract_SSD(file_path)
print("------ssd------")
print(res)
res2 = extract_gantry_angle(file_path)
print("------gantry------")
print(res2)



# 3. wedge
def extract_wedge(file_path):
    # step1: read one dicom file from given path and assigns value to the variable ds
    try:
        ds = pydicom.dcmread(file_path, force=True)
    except IOError:
        print("Error: The file was not found or failed to read")
    res = []
    # print(ds)
    # step2. find all wedge from ds
    # Number of Wedges
    # Beam Sequence

    try:
        for bs in ds.BeamSequence:
            if hasattr(bs, 'NumberOfWedges'):
                res.append(bs.NumberOfWedges)
    except AttributeError as err:
        print("OS error: {0}".format(err))

    # step3: Use 'set' to remove the same value
    res = list(set(res))
    return res


file_path = "/Users/yaozhiyuan/myunimelb/semster3/software project/DICOM/LII DICOM samples/YellowLvlIII_7a.dcm"
# res = extract_wedge(file_path)
# print(res)




def extract_isocentre(file_path):
    # step1: read one dicom file from given path and assigns value to the variable ds
    try:
        ds = pydicom.dcmread(file_path, force=True)
    except IOError:
        print("Error: The file was not found or failed to read")
    res = []
    # print(ds)
    # step2. find all gantry angles from ds
    # Isocenter Position
    # Control Point Sequence
    # Beam Sequence
    print(ds)
    try:
        for bs in ds.BeamSequence:
            if hasattr(bs, 'ControlPointSequence'):
                for cps in bs.ControlPointSequence:
                    if hasattr(cps, 'IsocenterPosition'):
                        res.append(cps.IsocenterPosition)
    except AttributeError as err:
        print("OS error: {0}".format(err))
    return res
    # step3: Use 'set' to remove the same value
    tmp =[]
    for r1 in res:
        r1.sort()
        for t in tmp:
            if r1 ==t:
                break
        else:
            tmp.append(r1)
    return tmp

file_path = "/Users/yaozhiyuan/myunimelb/semster3/software project/DICOM/LII DICOM samples/YellowLvlIII_7a.dcm"
# res = extract_isocentre(file_path)
# print(res)

#
#
# def extract_isocentre(file_path):
#     # step1: read one dicom file from given path and assigns value to the variable ds
#     try:
#         ds = pydicom.dcmread(file_path, force=True)
#     except IOError:
#         print("Error: The file was not found or failed to read")
#     res = []
#     print(ds)
#     # step2. find all gantry angles from ds
#     # Isocenter Position
#     # Control Point Sequence
#     # Beam Sequence
#     print(ds)
#     try:
#         for bs in ds.BeamSequence:
#             if hasattr(bs, 'ControlPointSequence'):
#                 for cps in bs.ControlPointSequence:
#                     if hasattr(cps, 'IsocenterPosition'):
#                         res.append(cps.IsocenterPosition)
#     except AttributeError as err:
#         print("OS error: {0}".format(err))
#
#     # step3: Use 'set' to remove the same value
#     # res = list(set(res))
#     return res

file_path = "/Users/yaozhiyuan/myunimelb/semster3/software project/DICOM/LII DICOM samples/YellowLvlIII_7a.dcm"
# res = extract_isocentre(file_path)
# print(res)


'''
validate collimator 
Beam Limiting Device Angle
'''
def extract_collimator(file_path):
    # step1: read one dicom file from given path and assigns value to the variable ds
    try:
        ds = pydicom.dcmread(file_path, force=True)
    except IOError:
        print("Error: The file was not found or failed to read")
    res = []
    # print(ds)
    # step2: find all gantry angles from ds
    # extract_collimator
    # Beam Limiting Device Angle
    # Control Point Sequence
    # Beam Sequence
    i =0
    try:
        for bs in ds.BeamSequence:
            if hasattr(bs, 'ControlPointSequence'):
                for cps in bs.ControlPointSequence:
                    if hasattr(cps, 'BeamLimitingDeviceAngle'):
                        res.append(cps.BeamLimitingDeviceAngle)
                        # print(str(i) +" :"+ str(cps.BeamLimitingDeviceAngle))
                        # i+=1
    except AttributeError as err:
        print("OS error: {0}".format(err))

    # step3: Use 'set' to remove the same value
    res = list(set(res))
    return res

# file_path = "/Users/yaozhiyuan/myunimelb/semster3/software project/DICOM/LII DICOM samples/YellowLvlIII_7a.dcm"
# res = extract_collimator(file_path)
# print(res)